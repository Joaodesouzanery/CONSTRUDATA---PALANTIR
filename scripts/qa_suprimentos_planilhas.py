#!/usr/bin/env python3
"""QA rapido para planilhas do modulo Suprimentos.

Uso:
  python scripts/qa_suprimentos_planilhas.py --materiais "f:/MATERIAIS_PENDENTES_v4.xlsx"

O script valida cabecalhos, separa materiais principais de auxiliares e soma
quantidades/metragens por nucleo, rua e rede. Ele e intencionalmente read-only.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def text(value: Any) -> str:
    return str(value or "").strip()


def number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    raw = re.sub(r"[^\d,.-]", "", text(value))
    if not raw:
        return 0.0
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def find_header(rows: list[list[Any]]) -> tuple[int, list[str]]:
    expected = {"material", "un", "rede", "qtd", "metragem", "nucleo", "rua"}
    for idx, row in enumerate(rows[:20]):
        headers = [text(cell).lower().replace("ú", "u") for cell in row]
        if len(expected.intersection(headers)) >= 4:
            return idx, headers
    raise ValueError("Cabecalho principal nao encontrado nas primeiras 20 linhas.")


def analyse_materiais(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    header_idx, headers = find_header(rows)
    col = {name: idx for idx, name in enumerate(headers)}

    totals: dict[str, dict[str, float]] = {}
    main_items = 0
    aux_items = 0
    current_nucleo = "Sem Nucleo"
    current_rua = "Sem Rua"

    for row in rows[header_idx + 1 :]:
        material = text(row[col.get("material", 0)] if col.get("material") is not None else "")
        if not material:
            continue
        nucleo = text(row[col["nucleo"]]) if "nucleo" in col else ""
        rua = text(row[col["rua"]]) if "rua" in col else ""
        if nucleo:
            current_nucleo = nucleo
        if rua:
            current_rua = rua

        rede = text(row[col["rede"]]).upper() if "rede" in col else ""
        qtd = number(row[col["qtd"]]) if "qtd" in col else 0.0
        metragem = number(row[col["metragem"]]) if "metragem" in col else 0.0
        is_aux = not rede or material.startswith(("•", "-", "  •"))
        if is_aux:
            aux_items += 1
        else:
            main_items += 1

        key = f"{current_nucleo} | {current_rua} | {rede or 'AUX'}"
        bucket = totals.setdefault(key, {"qtd": 0.0, "metragem_m": 0.0, "linhas": 0})
        bucket["qtd"] += qtd
        bucket["metragem_m"] += metragem
        bucket["linhas"] += 1

    return {
        "arquivo": str(path),
        "aba": ws.title,
        "linhas_lidas": len(rows) - header_idx - 1,
        "materiais_principais": main_items,
        "itens_auxiliares": aux_items,
        "agrupamentos": totals,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--materiais", type=Path, required=True)
    args = parser.parse_args()
    print(json.dumps(analyse_materiais(args.materiais), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
